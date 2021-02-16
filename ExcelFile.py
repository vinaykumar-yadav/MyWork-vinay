import openpyxl
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
wb=openpyxl.load_workbook(filename='demo.xlsx')
sheet=wb.active
print(sheet.cell(row=1,column=1).value)
'''blue_backg=PatternFill(bgColor=colors.BLUE)
diff_style=DifferentialStyle(fill=blue_backg)
rule=Rule(type='expression',dxf=diff_style)
rule.formula=["$A1<3"]
sheet.conditional_formatting.add("A1:C10",rule)'''
'''bold_font=Font(bold=True)
big_red_text=Font(color=colors.BLUE,size=20)
center_align=Alignment(horizontal='center')
border_side=Side(border_style='double')
sheet["C5"].font=bold_font
sheet["C6"].font=big_red_text
sheet["C7"].alignment=center_align
sheet["C8"].border=border_side'''
'''header=NamedStyle(name='header')
header.font=Font(bold=True)
header.border=Border(bottom=Side(border_style="thin"))
header.alignment=Alignment(horizontal="center",vertical="center")
header_row=sheet[4]
for cell in header_row:
    cell.style=header'''
#wb.save(filename='Java Freshers Schedule.xlsx')
from openpyxl.chart import BarChart,BarChart3D,AreaChart,LineChart, Reference
wb=openpyxl.Workbook()
sheet=wb.active
for i in range(10):
    sheet.append([input('Enter value')])
wb.save("demo.xlsx")
'''#values=Reference(sheet,min_col=1,min_row=1,max_col=1,max_row=10)
#chart=BarChart3D()
#chart=AreaChart()'''
'''chart=LineChart()
chart.add_data(values)
chart.title="BAR-CHART"
chart.x_axis.title="X-Axis"
chart.y_axis.title="Y-Axis"
sheet.add_chart(chart,"E2")
wb.save("Linechart.xlsx")'''